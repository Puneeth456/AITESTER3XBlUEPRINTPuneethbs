import os
import sys
import json
import re
from datetime import date, datetime
from pathlib import Path

from groq import Groq
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

BASE_DIR = Path(__file__).parent
LOG_FILE = BASE_DIR / "run_log.json"
EXCEL_FILE = BASE_DIR / "content_log.xlsx"

TOPICS = [
    "QA", "AI", "MCP", "RAG", "LLM", "AI Agents",
    "N8n", "Langflow", "CrewAI", "DeepEval",
    "LangChain", "AI Harness", "LLM Evaluation"
]


def get_today_topic():
    day_of_year = date.today().timetuple().tm_yday
    return TOPICS[day_of_year % len(TOPICS)]


def load_log():
    if LOG_FILE.exists():
        with open(LOG_FILE, encoding="utf-8-sig") as f:
            return json.load(f)
    return []


def save_log(log):
    with open(LOG_FILE, "w", encoding="utf-8") as f:
        json.dump(log, f, indent=2, ensure_ascii=False)


def is_already_logged(log, today_str):
    return any(entry["date"] == today_str for entry in log)


def parse_json_response(raw):
    raw = (raw or "").strip()
    # Always extract by outermost braces — avoids false-matching backticks
    # inside devto code snippets that the fence regex would misread.
    start = raw.find("{")
    end = raw.rfind("}")
    if start != -1 and end != -1 and end > start:
        raw = raw[start:end + 1]
    return json.loads(raw, strict=False)


def generate_content(topic, api_key):
    client = Groq(api_key=api_key)

    prompt = (
        f"You are a professional tech content creator. Today's topic is: {topic}\n\n"
        "Generate content for all 5 platforms. Return ONLY valid JSON with no markdown fences, "
        "no explanation, no extra text before or after. Use these exact keys:\n"
        '{\n'
        '  "linkedin_post": "professional post 150-300 words, end with 5-8 hashtags",\n'
        '  "medium_article": "full article 800-1200 words, title on line 1, use ## for subheadings",\n'
        '  "instagram_script": "bold hook on first line, caption under 150 words total, 3-5 hashtags",\n'
        '  "youtube_script": "hook intro + main body + subscribe CTA, 300-500 words",\n'
        '  "devto_article": "technical article 600-1000 words, ## sections, code snippets in triple backticks"\n'
        '}'
    )

    response = client.chat.completions.create(
        model="llama-3.3-70b-versatile",
        max_tokens=4000,
        messages=[{"role": "user", "content": prompt}]
    )
    return parse_json_response(response.choices[0].message.content)


def ensure_excel_headers(ws):
    headers = [
        "Date", "Topic", "LinkedIn Post", "Medium Article",
        "Instagram Script", "YouTube Script", "Dev.to Article", "Status"
    ]
    widths = [12, 16, 50, 80, 40, 50, 60, 10]
    fill = PatternFill("solid", fgColor="1E1E2E")
    font = Font(bold=True, color="CDD6F4", name="Calibri")

    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 20


def append_to_excel(today_str, topic, content, status):
    if EXCEL_FILE.exists():
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Content Log"
        ensure_excel_headers(ws)

    ws.append([
        today_str,
        topic,
        content.get("linkedin_post", ""),
        content.get("medium_article", ""),
        content.get("instagram_script", ""),
        content.get("youtube_script", ""),
        content.get("devto_article", ""),
        status,
    ])

    last_row = ws.max_row
    for col in range(3, 8):
        ws.cell(row=last_row, column=col).alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(EXCEL_FILE)


def run():
    today_str = date.today().isoformat()
    topic = get_today_topic()
    log = load_log()

    if is_already_logged(log, today_str):
        print(f"[OK] Already ran today ({today_str}): {topic} -- skipping")
        return

    api_key = os.environ.get("GROQ_API_KEY")
    if not api_key:
        print("[ERR] GROQ_API_KEY environment variable not set")
        sys.exit(1)

    print(f"[->] Date:  {today_str}")
    print(f"[->] Topic: {topic}")

    content = {}
    status = "error"

    try:
        print("[->] Calling Groq llama-3.3-70b-versatile (~10s)...")
        content = generate_content(topic, api_key)
        status = "success"
        print(f"[OK] Generated {len(content)} content pieces")
    except json.JSONDecodeError as e:
        print(f"[ERR] JSON parse failed: {e}")
    except Exception as e:
        print(f"[ERR] API error: {e}")

    try:
        append_to_excel(today_str, topic, content, status)
        print(f"[OK] Written to {EXCEL_FILE.name}")
    except Exception as e:
        print(f"[ERR] Excel write failed: {e}")

    log.append({
        "date": today_str,
        "topic": topic,
        "status": status,
        "timestamp": datetime.now().isoformat(),
        "content": content,
    })
    save_log(log)
    print(f"[OK] Logged to {LOG_FILE.name}")

    if status == "success":
        print(f"\n[OK] Done -- {topic} content ready in dashboard")
    else:
        print(f"\n[ERR] Run completed with errors -- check above")


if __name__ == "__main__":
    run()
